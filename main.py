import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de Catálogos Excel")
        self.root.geometry("900x700")
        
        self.df = None
        self.lineas_fijas = [1, 2, 8, 31, 32]
        
        self.setup_ui()
    
    # ... (Resto de setup_ui, seleccionar_todas, deseleccionar_todas, log)
    
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Botón para subir archivo
        self.btn_subir = ttk.Button(main_frame, text="Subir Archivo Excel", command=self.subir_archivo)
        self.btn_subir.grid(row=0, column=0, pady=10, sticky=tk.W)
        
        # Label para mostrar archivo seleccionado
        self.lbl_archivo = ttk.Label(main_frame, text="No se ha seleccionado archivo")
        self.lbl_archivo.grid(row=0, column=1, pady=10, sticky=tk.W)
        
        # Frame para selección de zona
        frame_zona = ttk.LabelFrame(main_frame, text="Seleccionar Zona", padding="10")
        frame_zona.grid(row=1, column=0, columnspan=2, pady=10, sticky=(tk.W, tk.E))
        
        self.zona_var = tk.StringVar(value="GBA-CABA")
        ttk.Radiobutton(frame_zona, text="GBA-CABA", variable=self.zona_var, value="GBA-CABA").grid(row=0, column=0, sticky=tk.W)
        ttk.Radiobutton(frame_zona, text="INTERIOR", variable=self.zona_var, value="INTERIOR").grid(row=0, column=1, sticky=tk.W)
        
        # Frame para checkboxes de líneas FIJAS
        self.frame_lineas = ttk.LabelFrame(main_frame, text="Seleccionar Líneas (Fijas)", padding="10")
        self.frame_lineas.grid(row=2, column=0, columnspan=2, pady=10, sticky=(tk.W, tk.E))
        
        # Botones para selección rápida
        frame_botones_lineas = ttk.Frame(self.frame_lineas)
        frame_botones_lineas.grid(row=0, column=0, columnspan=5, pady=5, sticky=(tk.W, tk.E))
        
        ttk.Button(frame_botones_lineas, text="Seleccionar Todas", 
                   command=self.seleccionar_todas).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_botones_lineas, text="Deseleccionar Todas", 
                   command=self.deseleccionar_todas).pack(side=tk.LEFT, padx=5)
        
        # Frame para los checkboxes de líneas FIJAS
        self.frame_checkboxes = ttk.Frame(self.frame_lineas)
        self.frame_checkboxes.grid(row=1, column=0, columnspan=5, pady=5, sticky=(tk.W, tk.E))
        
        # Crear checkboxes para las líneas fijas
        self.check_vars = {}
        for i, linea in enumerate(self.lineas_fijas):
            var = tk.BooleanVar()
            chk = ttk.Checkbutton(self.frame_checkboxes, text=f"Línea {linea}", variable=var)
            chk.grid(row=0, column=i, sticky=tk.W, padx=10, pady=2)
            self.check_vars[linea] = var
        
        # Botón para procesar
        self.btn_procesar = ttk.Button(main_frame, text="Procesar y Exportar EXCEL", command=self.procesar_archivo)
        self.btn_procesar.grid(row=3, column=0, pady=10, sticky=tk.W)
        self.btn_procesar.state(['disabled'])
        
        # Área de texto para logs
        self.text_log = tk.Text(main_frame, height=15, width=100)
        self.text_log.grid(row=4, column=0, columnspan=2, pady=10, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbar para el área de texto
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.text_log.yview)
        scrollbar.grid(row=4, column=2, pady=10, sticky=(tk.N, tk.S))
        self.text_log.configure(yscrollcommand=scrollbar.set)
        
        main_frame.rowconfigure(4, weight=1)
    
    def seleccionar_todas(self):
        for var in self.check_vars.values():
            var.set(True)
    
    def deseleccionar_todas(self):
        for var in self.check_vars.values():
            var.set(False)
    
    def log(self, message):
        self.text_log.insert(tk.END, message + "\n")
        self.text_log.see(tk.END)
        self.root.update()
    
    def subir_archivo(self):
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        
        if file_path:
            self.lbl_archivo.config(text=os.path.basename(file_path))
            try:
                columnas_a_texto = {'Codigo': str}
                # Modificación: Saltar las primeras 11 filas (1 a 11) y tomar la fila 12 (índice 11) como encabezado
                self.df = pd.read_excel(file_path, header=11, dtype=columnas_a_texto)
                
                self.log(f"Archivo cargado: {os.path.basename(file_path)}")
                self.log(f"Filas cargadas (después de saltar el inicio): {len(self.df)}")
                self.log(f"Columnas (tomadas de la fila 12 original): {list(self.df.columns)}")
                
                self.btn_procesar.state(['!disabled'])
                
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo cargar el archivo: {str(e)}")
                self.log(f"ERROR: {str(e)}")
    
    def extraer_numero_rubro(self, rubro):
        if pd.isna(rubro) or rubro == "":
            return 999
        
        rubro_str = str(rubro)
        match = re.match(r'^(\d+)', rubro_str)
        if match:
            return int(match.group(1))
        else:
            return 999
    
    def procesar_archivo(self):
        if self.df is None:
            messagebox.showwarning("Advertencia", "Primero debe cargar un archivo Excel")
            return
        
        lineas_seleccionadas = [linea for linea in self.lineas_fijas if self.check_vars[linea].get()]
        
        if not lineas_seleccionadas:
            messagebox.showwarning("Advertencia", "Debe seleccionar al menos una línea")
            return
        
        zona = self.zona_var.get()
        self.log(f"Procesando líneas: {lineas_seleccionadas} - Zona: {zona}")
        
        try:
            # Paso 1: Filtrar por líneas seleccionadas
            columna_linea = self.identificar_columna_linea()
            if columna_linea:
                df_filtrado = self.df[self.df[columna_linea].isin(lineas_seleccionadas)].copy()
                self.log(f"Filas después de filtrar por líneas: {len(df_filtrado)}")
            else:
                self.log("ADVERTENCIA: No se encontró columna de líneas, procesando todo el archivo")
                df_filtrado = self.df.copy()
            
            # Paso 2: Aplicar reglas de columnas según la zona
            df_filtrado = self.aplicar_reglas_columnas(df_filtrado, zona)
            
            # Paso 3: Aplicar reglas de selección de precios
            df_filtrado = self.aplicar_reglas_precio(df_filtrado, zona)
            
            # Paso 4: Ordenar por Rubro (numérico) y Marca (alfabético)
            if 'Rubro' in df_filtrado.columns and 'Marca' in df_filtrado.columns:
                df_filtrado['rubro_numero'] = df_filtrado['Rubro'].apply(self.extraer_numero_rubro)
                # Ordenamiento solo por Rubro (numérico) y luego por Marca (alfabético)
                df_filtrado = df_filtrado.sort_values(['rubro_numero', 'Marca'], ascending=[True, True])
                df_filtrado = df_filtrado.drop('rubro_numero', axis=1)
                self.log("Datos ordenados por Rubro (numérico) y Marca (alfabético)")
            
            # Paso 5: Renumerar orden desde 1 (SIN importar la fila de inicio del catálogo)
            df_filtrado = df_filtrado.reset_index(drop=True)
            df_filtrado['orden'] = range(1, len(df_filtrado) + 1)
            self.log(f"Orden renumerado del 1 al {len(df_filtrado)}")
            
            # Paso 6: Aplicar regla del múltiplo de 8
            df_final = self.aplicar_multiplo_8(df_filtrado)
            
            # Paso 7: Exportar a Excel con formato
            self.exportar_excel_con_formato(df_final, zona)
            
            messagebox.showinfo("Éxito", "Archivo procesado y exportado correctamente a Excel")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error durante el procesamiento: {str(e)}")
            self.log(f"ERROR: {str(e)}")
    
    def identificar_columna_linea(self):
        posibles_columnas = ['Linea', 'linea', 'LINEA', 'Línea']
        for col in posibles_columnas:
            if col in self.df.columns:
                return col
        # Si no la encuentra, intenta identificarla a partir de las columnas cargadas
        # El archivo de ejemplo tiene la columna de línea como el primer campo sin nombre después de 'orden'
        if len(self.df.columns) > 1 and str(self.df.columns[1]).lower() in ['unnamed: 1', 'unnamed: 0']:
            return self.df.columns[1] # Esto es una suposición basada en el formato habitual sin encabezado
        return None
    
    # ... (aplicar_reglas_columnas, aplicar_reglas_precio, aplicar_multiplo_8, crear_filas_vacias)
    
    def aplicar_reglas_columnas(self, df, zona):
        self.log(f"Aplicando reglas de columnas para zona: {zona}")
        
        if zona == "INTERIOR":
            columnas_a_eliminar = []
            columnas = list(df.columns)
            inicio_eliminar = None
            fin_eliminar = None
            
            # Buscar columnas que contienen 'lista1' y 'ad 3' (o nombres similares)
            for i, col in enumerate(columnas):
                col_str = str(col).lower()
                if 'lista1' in col_str and inicio_eliminar is None:
                    inicio_eliminar = i
                if 'ad 3' in col_str and inicio_eliminar is not None:
                    fin_eliminar = i
                    break
            
            if inicio_eliminar is not None and fin_eliminar is not None:
                columnas_a_eliminar = columnas[inicio_eliminar:fin_eliminar + 1]
                df = df.drop(columns=columnas_a_eliminar)
                self.log(f"Columnas eliminadas para INTERIOR: {columnas_a_eliminar}")
        
        return df
    
    def aplicar_reglas_precio(self, df, zona):
        self.log("Aplicando reglas de precios...")
        
        # Nombres de columnas de precio (deben coincidir con el encabezado de la fila 12)
        columnas_precio = {
            'GBA-CABA': {'precio_default': 'l1 5', 'precio_oferta': 'l1 9'},
            'INTERIOR': {'precio_default': 'l2 5', 'precio_oferta': 'l2 9'}
        }
        
        col_default = columnas_precio[zona]['precio_default']
        col_oferta = columnas_precio[zona]['precio_oferta']
        
        # Intenta encontrar las columnas de precio, si no están, usa alternativas
        if col_default not in df.columns or col_oferta not in df.columns:
            self.log(f"ADVERTENCIA: No se encontraron columnas de precio específicas para {zona} ({col_default}, {col_oferta})")
            if 'l1 5' in df.columns and 'l1 9' in df.columns:
                col_default, col_oferta = 'l1 5', 'l1 9'
                self.log(f"Usando columnas alternativas: {col_default}, {col_oferta}")
            else:
                self.log("ERROR: No se encontraron columnas de precio adecuadas")
                return df
        
        # Asegurar que las columnas de precio sean numéricas para evitar errores en la comparación
        df[col_default] = pd.to_numeric(df[col_default], errors='coerce')
        df[col_oferta] = pd.to_numeric(df[col_oferta], errors='coerce')

        df['precio_seleccionado'] = df[col_default]
        
        # La lógica de oferta usa las columnas 'orden' (que debe ser 'ord' según el archivo de ejemplo) y 'condicion'
        # Ajusto 'orden' por 'ord' si está disponible
        col_ord = 'ord' if 'ord' in df.columns else 'orden' 

        mascara_oferta = (
            df[col_ord].astype(str).str.contains(r'a\d{2}\.eps', na=False) & 
            df['condicion'].astype(str).str.contains('oferta', na=False, case=False)
        )
        
        df.loc[mascara_oferta, 'precio_seleccionado'] = df.loc[mascara_oferta, col_oferta]
        
        self.log(f"Reglas de precios aplicadas. Ofertas encontradas: {mascara_oferta.sum()}")
        
        return df

    def aplicar_multiplo_8(self, df):
        self.log("Aplicando regla del múltiplo de 8...")
        
        if 'Rubro' not in df.columns:
            self.log("ADVERTENCIA: No hay columna 'Rubro', no se aplicará múltiplo de 8")
            return df
        
        dfs_por_rubro = []
        # El contador de orden debe continuar a partir del último 'orden' existente
        contador_orden = df['orden'].max() + 1 if 'orden' in df.columns and not df['orden'].empty else 1
        
        for rubro, grupo in df.groupby('Rubro'):
            filas_rubro = len(grupo)
            filas_necesarias = ((filas_rubro + 7) // 8) * 8
            filas_faltantes = filas_necesarias - filas_rubro
            
            self.log(f"Rubro {rubro}: {filas_rubro} filas, necesarias: {filas_necesarias}, faltantes: {filas_faltantes}")
            
            if filas_faltantes > 0:
                filas_vacias = self.crear_filas_vacias(filas_faltantes, rubro, grupo, contador_orden)
                contador_orden += filas_faltantes
                grupo = pd.concat([grupo, filas_vacias], ignore_index=True)
            
            dfs_por_rubro.append(grupo)
        
        df_final = pd.concat(dfs_por_rubro, ignore_index=True)
        # Renumerar el 'orden' final después de las filas vacías
        df_final['orden'] = range(1, len(df_final) + 1) 
        self.log(f"Total filas después de aplicar múltiplo de 8: {len(df_final)}")
        
        return df_final
    
    def crear_filas_vacias(self, cantidad, rubro, grupo_referencia, inicio_orden):
        if len(grupo_referencia) == 0:
            return pd.DataFrame()
        
        primera_fila = grupo_referencia.iloc[0]
        
        filas_vacias = []
        for i in range(cantidad):
            fila_vacia = {}
            fila_vacia['orden'] = inicio_orden + i
            # Asegurarse de rellenar todas las columnas existentes con vacío o valor por defecto
            for col in grupo_referencia.columns:
                if col in ['orden']:
                    continue
                elif col in ['ord']:
                    fila_vacia[col] = "Vacio.eps"
                elif col in ['condicion', 'Marca', 'Codigo', 'Descripcion', 'peso']:
                    fila_vacia[col] = ""
                elif col == 'Rubro':
                    fila_vacia[col] = rubro
                elif col in ['0.05', '0.07', '9/especial', '0.11']: # Ejemplo de columnas con valores por defecto
                    fila_vacia[col] = 0.00 
                elif col == 'precio_seleccionado':
                    fila_vacia[col] = np.nan # Usar NaN para que quede vacío o 0 en Excel
                else:
                    fila_vacia[col] = ""
            
            filas_vacias.append(fila_vacia)
        
        return pd.DataFrame(filas_vacias)
    
    def aplicar_formato_numeros_excel(self, df):
        """Aplica formato de números para Excel (limpieza de datos)"""
        df_formateado = df.copy()
        
        # Columnas que deben ser numéricas para el cálculo y openpyxl
        columnas_numericas = ['precio_seleccionado', 'lista1', 'lista 2', 'lista 3', 'lista 4', 'lista 5',
                                    'l1 5', 'l1 7',  'l1 9', 'l1 11',
                                    'l2 5', 'l2 7', 'l2 9',  'l2 11',
                                    'l3 5', 'l3 7', 'l3 9',  'l3 11',
                                    'l4 5', 'l4 7', 'l4 9',  'l4 11',
                                    'l2 5.1', 'l2 7.1', 'l2 9.1', 'l2 11.1',
                                    '0.05', '0.07', '9/especial', '0.11']
        
        for col in columnas_numericas:
            if col in df_formateado.columns:
                # 1. Convertir a numérico forzando errores a NaN
                df_formateado[col] = pd.to_numeric(df_formateado[col], errors='coerce')
                
                # 2. **Redondear a 2 decimales en el DataFrame**
                df_formateado[col] = df_formateado[col].round(2)
        
        return df_formateado
    
    def exportar_excel_con_formato(self, df, zona):
        """Exporta a Excel con formato profesional"""
        file_path = filedialog.asksaveasfilename(
            title="Guardar archivo Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if file_path:
            try:
                # Aplicar formato de números (limpieza de datos)
                df_export = self.aplicar_formato_numeros_excel(df)
                
                # Crear libro de Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Catálogo Procesado"
                
                # Agregar encabezados informativos (Filas 1 a 5)
                ws['A1'] = "CATÁLOGO MADRE - EXPORTACIÓN"
                ws['A2'] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A3'] = f"Zona: {zona}"
                ws['A4'] = f"Líneas procesadas: {[linea for linea in self.lineas_fijas if self.check_vars[linea].get()]}"
                ws['A5'] = f"Total productos: {len(df)}"
                
                # Agregar datos (Empieza en Fila 6, Header en Fila 6, Datos en Fila 7)
                for r_idx, r in enumerate(dataframe_to_rows(df_export, index=False, header=True)):
                    ws.append(r)
                
                # Aplicar formato a las celdas (El encabezado de datos está en la Fila 6)
                self.aplicar_estilos_excel(ws, len(df_export), start_row_data=7, header_row=6)
                
                # Guardar archivo
                wb.save(file_path)
                
                self.log(f"Archivo Excel exportado: {file_path}")
                self.log(f"Total de productos: {len(df)}")
                self.log("Formato Excel aplicado: números con 2 decimales, estilos profesionales")
                
                # Mostrar preview
                self.mostrar_preview_excel(df_export)
                
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo exportar el archivo Excel: {str(e)}")
                self.log(f"ERROR en exportación: {str(e)}")
    
    def aplicar_estilos_excel(self, ws, total_filas, start_row_data, header_row):
        """Aplica estilos profesionales al Excel, incluyendo formato de 2 decimales robusto."""
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Identificar las columnas de precios por el nombre del header
        header_cell_values = [cell.value for cell in ws[header_row]]
        precio_cols = {}
        columnas_numericas_clave = ['precio_seleccionado', 'lista1', 'lista 2', 'lista 3', 'lista 4', 'lista 5',
                                    'l1 5', 'l1 7',  'l1 9', 'l1 11',
                                    'l2 5', 'l2 7', 'l2 9',  'l2 11',
                                    'l3 5', 'l3 7', 'l3 9',  'l3 11',
                                    'l4 5', 'l4 7', 'l4 9',  'l4 11',
                                    'l2 5.1', 'l2 7.1', 'l2 9.1', 'l2 11.1',
                                    '0.05', '0.07', '9/especial', '0.11']
        
        for idx, col_name in enumerate(header_cell_values):
            col_letter = get_column_letter(idx + 1)
            if col_name in columnas_numericas_clave:
                precio_cols[col_letter] = col_name

        # Aplicar formato a encabezados
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Aplicar formato a datos numéricos (Fila 7 en adelante)
        for row in range(start_row_data, total_filas + start_row_data):
            for cell in ws[row]:
                cell.border = border
                
                # Aplicar formato numérico: Separador de miles y dos decimales
                if cell.column_letter in precio_cols:
                    # Usar el formato de Excel para miles y dos decimales: 0.00
                    cell.number_format = '0.00' 
        
        # Autoajustar columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def mostrar_preview_excel(self, df):
        """Muestra un preview de los datos que se exportarán a Excel"""
        self.log("\n--- PREVIEW PARA EXCEL ---")
        self.log(f"Primeras 3 filas con formato:")
        
        columnas_preview = ['orden', 'Rubro', 'Marca', 'precio_seleccionado']
        columnas_disponibles = [col for col in columnas_preview if col in df.columns]
        
        if columnas_disponibles:
            preview_df = df[columnas_disponibles].head(3).copy()
            
            # Formatear números para preview (solo para el log de texto)
            if 'precio_seleccionado' in preview_df.columns:
                preview_df['precio_seleccionado'] = preview_df['precio_seleccionado'].apply(
                    lambda x: f"{x:,.2f}" if pd.notna(x) and x != "" else ""
                )
            
            self.log(preview_df.to_string(index=False))
        else:
            self.log(df.head(3).to_string(index=False))
            
        self.log("--- FIN PREVIEW ---\n")
        self.log("✅ Archivo listo para exportar a Excel")
        self.log("💡 Consejo: Revise el Excel y luego exporte a TXT tabulado manualmente")

def main():
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()